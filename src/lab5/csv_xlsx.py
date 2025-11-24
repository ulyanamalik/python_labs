import csv  # Импорт библиотеки для работы с CSV файлами
from pathlib import Path  # Импорт модуля для удобной работы с путями файлов
from openpyxl import Workbook  # Импорт основной библиотеки для работы с Excel файлами
from openpyxl.utils import (
    get_column_letter,
)  # Импорт утилиты для перевода индекса столбца в букву Excel


# Функция для проверки типа файла по расширению
def check_file_type(file_path: str, valid_types: tuple) -> bool:
    """
    Проверяет, соответствует ли расширение файла одному из указанных типов.
    Возвращает True, если файл имеет одно из разрешенных расширений, иначе False.
    """
    return Path(file_path).suffix.lower() in valid_types


# Функция для преобразования CSV файла в XLSX файл
def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """

    # Проверка расширения входящего файла (должен быть .csv)
    if not check_file_type(csv_path, (".csv",)):
        raise ValueError(f"Входной файл '{csv_path}' не является CSV.")

    # Проверка расширения выходящего файла (должен быть .xlsx)
    if not check_file_type(xlsx_path, (".xlsx",)):
        raise ValueError(f"Выходной файл '{xlsx_path}' не является XLSX.")

    # Создание объекта пути для CSV файла
    csv_file = Path(csv_path)

    # Проверка существования CSV файла
    if not csv_file.exists():
        # Поднимается исключение, если файл не найден
        raise FileNotFoundError(f"Файл {csv_path} не найден.")  # [1]

    # Открытие CSV файла для чтения
    with csv_file.open("r", encoding="utf-8") as f:
        # Создается объект reader для парсинга CSV
        reader = csv.reader(f)  # [2]
        # Читаются все строки CSV файла и превращаются в список списков
        data = list(reader)  # [3]

    # Проверка, есть ли данные в CSV файле
    if not data:
        # Поднимает ошибку, если файл пуст
        raise ValueError("CSV-файл пуст.")  # [4]

    # Создание нового рабочего документа Excel
    workbook = Workbook()  # [5]
    # Получение активного листа
    sheet = workbook.active  # [6]
    # Переименование листа в 'Sheet1'
    sheet.title = "Sheet1"  # [7]

    # Запись данных из CSV в лист Excel
    for row in data:
        # Каждая строка добавляется в лист
        sheet.append(row)  # [8]

    # Настройка ширины столбцов автоматически
    for col_idx, col in enumerate(sheet.columns, 1):
        # Для каждого столбца вычисляется максимальная длина значений
        max_length = max(len(str(cell.value)) for cell in col)  # [9]
        # Устанавливается минимальная ширина столбца (8 символов)
        adjusted_width = max((max_length + 2), 8)  # [10]
        # Применение ширины столбцу по индексу
        sheet.column_dimensions[get_column_letter(col_idx)].width = (
            adjusted_width  # [11]
        )

    # Сохранение Excel файла
    workbook.save(xlsx_path)  # [12]


# Основной код программы
if __name__ == "__main__":
    # Примеры использования функции
    csv_to_xlsx("data/samples/people.csv", "data/out/people.xlsx")  # [13]
